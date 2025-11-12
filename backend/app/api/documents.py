"""
ARIA ERP - Document Processing API
Handles document upload, OCR, classification, and posting to ERP
"""
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Form
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import datetime, date
from decimal import Decimal
import os
import hashlib
import shutil
from pathlib import Path
import io
import json

from core.database import get_db
from app.auth import get_current_user
from app.models.user import User

from bots.document_classification_bot import DocumentClassificationBot
from bots.ocr_invoice_bot import OCRInvoiceBot
from bots.bot_api_client import BotAPIClient

router = APIRouter(prefix="/api/documents", tags=["documents"])

UPLOAD_DIR = Path("/var/www/aria/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png'}


def get_file_hash(file_path: str) -> str:
    """Calculate SHA256 hash of file"""
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


@router.post("/process")
async def process_document(
    file: UploadFile = File(...),
    company_id: int = Form(...),
    vendor_id: Optional[int] = Form(None),
    doc_type_hint: Optional[str] = Form(None),
    instruction: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Process uploaded document: classify, run OCR, extract structured data
    
    Returns:
        - doc_type: Detected document type (invoice, receipt, statement, etc.)
        - header: Extracted header fields (vendor, dates, totals, VAT)
        - lines: Extracted line items
        - confidence: Confidence scores per field
        - file_id: Unique file identifier for later posting
        - warnings: Any issues detected (missing vendor, duplicate, etc.)
    """
    try:
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"File type not allowed. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
            )
        
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Maximum size: {MAX_FILE_SIZE / 1024 / 1024}MB"
            )
        
        file_hash = hashlib.sha256(contents).hexdigest()
        
        company_dir = UPLOAD_DIR / str(company_id)
        company_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = f"{timestamp}_{file_hash[:8]}{file_ext}"
        file_path = company_dir / safe_filename
        
        with open(file_path, "wb") as f:
            f.write(contents)
        
        classification_bot = DocumentClassificationBot()
        doc_classification = classification_bot.classify_document(str(file_path))
        
        doc_type = doc_type_hint or doc_classification.get('category', 'unknown')
        confidence = doc_classification.get('confidence', 0.0)
        sap_posting = doc_classification.get('sap_posting', {})
        
        ocr_data = {}
        line_items = []
        header_fields = {}
        field_confidence = {}
        
        if doc_type == 'invoice':
            ocr_bot = OCRInvoiceBot()
            ocr_data = ocr_bot.extract_invoice_data(str(file_path))
            
            header_fields = {
                'supplier_name': ocr_data.get('supplier_name', ''),
                'invoice_number': ocr_data.get('invoice_number', ''),
                'invoice_date': ocr_data.get('invoice_date', ''),
                'due_date': ocr_data.get('due_date', ''),
                'total_amount': float(ocr_data.get('total_amount', 0)),
                'vat_amount': float(ocr_data.get('vat_amount', 0)),
                'net_amount': float(ocr_data.get('total_amount', 0)) - float(ocr_data.get('vat_amount', 0))
            }
            
            line_items = ocr_data.get('line_items', [])
            
            overall_confidence = ocr_data.get('confidence_score', 0.95)
            field_confidence = {
                'supplier_name': overall_confidence,
                'invoice_number': overall_confidence,
                'invoice_date': overall_confidence,
                'due_date': overall_confidence,
                'total_amount': overall_confidence,
                'vat_amount': overall_confidence
            }
        
        warnings = []
        
        if not vendor_id and doc_type == 'invoice':
            warnings.append({
                'type': 'missing_vendor',
                'message': 'Vendor not specified. Please select a vendor before posting.',
                'severity': 'error'
            })
        
        if doc_type == 'invoice' and vendor_id and header_fields.get('invoice_number'):
            from app.models.ap import VendorBill
            existing_bill = db.query(VendorBill).filter(
                VendorBill.company_id == company_id,
                VendorBill.vendor_id == vendor_id,
                VendorBill.vendor_invoice_number == header_fields['invoice_number']
            ).first()
            
            if existing_bill:
                warnings.append({
                    'type': 'possible_duplicate',
                    'message': f'Possible duplicate: Bill #{existing_bill.bill_number} with same invoice number exists',
                    'severity': 'warning',
                    'existing_bill_id': existing_bill.id
                })
        
        sum_lines = sum(float(line.get('total', 0)) for line in line_items)
        if line_items and abs(sum_lines - header_fields.get('net_amount', 0)) > 0.01:
            warnings.append({
                'type': 'total_mismatch',
                'message': f'Line items total ({sum_lines:.2f}) does not match net amount ({header_fields.get("net_amount", 0):.2f})',
                'severity': 'warning'
            })
        
        return {
            'success': True,
            'file_id': file_hash[:16],
            'file_path': str(file_path),
            'doc_type': doc_type,
            'classification_confidence': confidence,
            'sap_posting': sap_posting,
            'header': header_fields,
            'lines': line_items,
            'field_confidence': field_confidence,
            'warnings': warnings,
            'instruction': instruction,
            'suggested_actions': ['post_to_ap', 'send_to_controller', 'save_draft']
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/post")
async def post_document(
    file_id: str = Form(...),
    file_path: str = Form(...),
    company_id: int = Form(...),
    vendor_id: int = Form(...),
    doc_type: str = Form(...),
    invoice_number: str = Form(...),
    invoice_date: str = Form(...),
    due_date: str = Form(...),
    total_amount: float = Form(...),
    vat_amount: float = Form(...),
    lines: str = Form(...),  # JSON string
    override_duplicate: bool = Form(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Post processed document to ERP (create AP bill with attachment)
    
    Args:
        file_id: Unique file identifier from process endpoint
        file_path: Path to uploaded file
        company_id: Company ID
        vendor_id: Vendor ID
        doc_type: Document type (invoice, receipt, etc.)
        invoice_number: Vendor invoice number
        invoice_date: Invoice date
        due_date: Due date
        total_amount: Total amount including VAT
        vat_amount: VAT amount
        lines: JSON string of line items
        override_duplicate: Override duplicate check
    
    Returns:
        - bill_id: Created bill ID
        - bill_number: Generated bill number
        - status: Posting status
    """
    try:
        import json
        from app.models.ap import VendorBill, VendorBillLine
        
        if not override_duplicate:
            existing_bill = db.query(VendorBill).filter(
                VendorBill.company_id == company_id,
                VendorBill.vendor_id == vendor_id,
                VendorBill.vendor_invoice_number == invoice_number
            ).first()
            
            if existing_bill:
                raise HTTPException(
                    status_code=409,
                    detail=f"Duplicate bill detected: Bill #{existing_bill.bill_number} already exists with this invoice number"
                )
        
        line_items = json.loads(lines) if isinstance(lines, str) else lines
        
        last_bill = db.query(VendorBill).filter(
            VendorBill.company_id == company_id
        ).order_by(VendorBill.id.desc()).first()
        
        next_number = 1
        if last_bill and last_bill.bill_number:
            try:
                next_number = int(last_bill.bill_number.split('-')[-1]) + 1
            except:
                next_number = 1
        
        bill_number = f"BILL-{next_number:05d}"
        
        net_amount = total_amount - vat_amount
        
        new_bill = VendorBill(
            company_id=company_id,
            vendor_id=vendor_id,
            bill_number=bill_number,
            vendor_invoice_number=invoice_number,
            bill_date=datetime.strptime(invoice_date, '%Y-%m-%d').date(),
            due_date=datetime.strptime(due_date, '%Y-%m-%d').date(),
            subtotal=Decimal(str(net_amount)),
            tax_amount=Decimal(str(vat_amount)),
            total_amount=Decimal(str(total_amount)),
            status='draft',
            created_by=current_user.id,
            created_at=datetime.now()
        )
        
        db.add(new_bill)
        db.flush()
        
        for idx, line in enumerate(line_items, 1):
            bill_line = VendorBillLine(
                bill_id=new_bill.id,
                line_number=idx,
                description=line.get('description', ''),
                quantity=Decimal(str(line.get('quantity', 1))),
                unit_price=Decimal(str(line.get('unit_price', 0))),
                discount_percent=Decimal(str(line.get('discount_percent', 0))),
                tax_rate=Decimal(str(line.get('tax_rate', 15))),
                line_total=Decimal(str(line.get('total', 0)))
            )
            db.add(bill_line)
        
        db.commit()
        db.refresh(new_bill)
        
        return {
            'success': True,
            'bill_id': new_bill.id,
            'bill_number': new_bill.bill_number,
            'status': new_bill.status,
            'message': f'Bill {new_bill.bill_number} created successfully',
            'file_id': file_id,
            'file_path': file_path
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/send-to-controller")
async def send_to_controller(
    file_id: str = Form(...),
    file_path: str = Form(...),
    instruction: str = Form(...),
    company_id: int = Form(...),
    current_user: User = Depends(get_current_user)
):
    """
    Send document to Aria Controller for NL-driven workflow
    
    This allows the controller to handle ambiguous cases or complex workflows
    that require multi-bot orchestration or user clarification.
    """
    return {
        'success': True,
        'message': 'Document sent to Aria Controller',
        'file_id': file_id,
        'instruction': instruction,
        'status': 'queued',
        'note': 'Controller integration coming soon - will process via email/workflow engine'
    }


@router.post("/export-excel")
async def export_to_excel(
    doc_type: str = Form(...),
    header: str = Form(...),
    lines: str = Form(...),
    sap_posting: str = Form(None),
    current_user: User = Depends(get_current_user)
):
    """
    Export extracted document data to Excel for manual SAP upload
    
    Creates a formatted Excel file with:
    - Header sheet with document details
    - Line items sheet with all line details
    - SAP posting instructions
    - Ready for manual upload to SAP
    
    Args:
        doc_type: Document type (invoice, credit_note, etc.)
        header: JSON string of header fields
        lines: JSON string of line items
        sap_posting: JSON string of SAP posting suggestions
    
    Returns:
        Excel file download
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_data = json.loads(header) if isinstance(header, str) else header
        line_items = json.loads(lines) if isinstance(lines, str) else lines
        sap_data = json.loads(sap_posting) if sap_posting and isinstance(sap_posting, str) else {}
        
        wb = Workbook()
        
        ws_header = wb.active
        ws_header.title = "Document Header"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_header['A1'] = 'ARIA ERP - Document Export for SAP'
        ws_header['A1'].font = Font(bold=True, size=14, color="4472C4")
        ws_header.merge_cells('A1:D1')
        
        ws_header['A2'] = f'Document Type: {doc_type.upper().replace("_", " ")}'
        ws_header['A2'].font = Font(bold=True, size=11)
        ws_header.merge_cells('A2:D2')
        
        ws_header['A3'] = f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_header['A3'].font = Font(italic=True, size=10)
        ws_header.merge_cells('A3:D3')
        
        if sap_data:
            row = 5
            ws_header[f'A{row}'] = 'SAP Posting Instructions'
            ws_header[f'A{row}'].font = Font(bold=True, size=12, color="C00000")
            ws_header.merge_cells(f'A{row}:D{row}')
            
            row += 1
            ws_header[f'A{row}'] = 'SAP Module:'
            ws_header[f'A{row}'].font = Font(bold=True)
            ws_header[f'B{row}'] = sap_data.get('module', 'N/A')
            ws_header[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            row += 1
            ws_header[f'A{row}'] = 'Transaction Code:'
            ws_header[f'A{row}'].font = Font(bold=True)
            ws_header[f'B{row}'] = sap_data.get('tcode', 'N/A')
            ws_header[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws_header[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
            
            row += 1
            ws_header[f'A{row}'] = 'Description:'
            ws_header[f'A{row}'].font = Font(bold=True)
            ws_header[f'B{row}'] = sap_data.get('description', 'N/A')
            ws_header.merge_cells(f'B{row}:D{row}')
            
            row += 1
            ws_header[f'A{row}'] = 'Rationale:'
            ws_header[f'A{row}'].font = Font(bold=True)
            ws_header[f'B{row}'] = sap_data.get('rationale', 'N/A')
            ws_header.merge_cells(f'B{row}:D{row}')
            ws_header[f'B{row}'].alignment = Alignment(wrap_text=True)
            
            row += 2
        else:
            row = 7
        
        ws_header[f'A{row}'] = 'Document Details'
        ws_header[f'A{row}'].font = Font(bold=True, size=12)
        ws_header[f'A{row}'].fill = header_fill
        ws_header[f'B{row}'].fill = header_fill
        ws_header.merge_cells(f'A{row}:B{row}')
        
        row += 1
        header_fields = [
            ('Supplier/Vendor', header_data.get('supplier_name', '')),
            ('Invoice Number', header_data.get('invoice_number', '')),
            ('Invoice Date', header_data.get('invoice_date', '')),
            ('Due Date', header_data.get('due_date', '')),
            ('Net Amount', header_data.get('net_amount', 0)),
            ('VAT Amount', header_data.get('vat_amount', 0)),
            ('Total Amount', header_data.get('total_amount', 0))
        ]
        
        for field_name, field_value in header_fields:
            ws_header[f'A{row}'] = field_name
            ws_header[f'A{row}'].font = Font(bold=True)
            ws_header[f'A{row}'].border = border
            ws_header[f'B{row}'] = field_value
            ws_header[f'B{row}'].border = border
            if isinstance(field_value, (int, float)):
                ws_header[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        ws_header.column_dimensions['A'].width = 25
        ws_header.column_dimensions['B'].width = 30
        ws_header.column_dimensions['C'].width = 20
        ws_header.column_dimensions['D'].width = 20
        
        if line_items:
            ws_lines = wb.create_sheet("Line Items")
            
            headers = ['Line #', 'Description', 'Quantity', 'Unit Price', 'Discount %', 'Tax Rate %', 'Line Total']
            for col_num, header_text in enumerate(headers, 1):
                cell = ws_lines.cell(row=1, column=col_num)
                cell.value = header_text
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for row_num, line in enumerate(line_items, 2):
                ws_lines.cell(row=row_num, column=1, value=row_num - 1).border = border
                ws_lines.cell(row=row_num, column=2, value=line.get('description', '')).border = border
                
                qty_cell = ws_lines.cell(row=row_num, column=3, value=float(line.get('quantity', 0)))
                qty_cell.number_format = '#,##0.00'
                qty_cell.border = border
                
                price_cell = ws_lines.cell(row=row_num, column=4, value=float(line.get('unit_price', 0)))
                price_cell.number_format = '#,##0.00'
                price_cell.border = border
                
                disc_cell = ws_lines.cell(row=row_num, column=5, value=float(line.get('discount_percent', 0)))
                disc_cell.number_format = '0.00'
                disc_cell.border = border
                
                tax_cell = ws_lines.cell(row=row_num, column=6, value=float(line.get('tax_rate', 15)))
                tax_cell.number_format = '0.00'
                tax_cell.border = border
                
                total_cell = ws_lines.cell(row=row_num, column=7, value=float(line.get('total', 0)))
                total_cell.number_format = '#,##0.00'
                total_cell.border = border
            
            ws_lines.column_dimensions['A'].width = 10
            ws_lines.column_dimensions['B'].width = 50
            ws_lines.column_dimensions['C'].width = 12
            ws_lines.column_dimensions['D'].width = 15
            ws_lines.column_dimensions['E'].width = 12
            ws_lines.column_dimensions['F'].width = 12
            ws_lines.column_dimensions['G'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ARIA_SAP_Export_{doc_type}_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")
