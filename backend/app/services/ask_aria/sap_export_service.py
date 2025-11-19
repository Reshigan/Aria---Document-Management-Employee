"""
SAP Export Service
Generates Excel/CSV templates for SAP document upload
"""
import os
import json
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class SAPExportService:
    """Service for generating SAP export templates"""
    
    def __init__(self, templates_path: Optional[str] = None):
        if templates_path is None:
            templates_path = Path(__file__).parent.parent / "templates" / "sap_templates_enhanced.json"
        
        self.templates_path = templates_path
        self.templates: List[Dict] = []
        self.templates_by_id: Dict[str, Dict] = {}
        self.templates_by_type: Dict[str, Dict] = {}
        
        self.load_templates()
    
    def load_templates(self):
        """Load SAP templates from JSON file"""
        try:
            with open(self.templates_path, 'r') as f:
                self.templates = json.load(f)
            
            for template in self.templates:
                self.templates_by_id[template['id']] = template
                self.templates_by_type[template['doc_type']] = template
            
            logger.info(f"Loaded {len(self.templates)} SAP export templates")
        except Exception as e:
            logger.error(f"Failed to load SAP templates: {e}")
            self.templates = []
    
    def generate_excel_export(
        self,
        doc_type: str,
        header_data: Dict[str, Any],
        line_items: List[Dict[str, Any]],
        output_path: Optional[str] = None
    ) -> bytes:
        """
        Generate Excel export file for SAP upload
        
        Args:
            doc_type: SAP document type (e.g., 'AP_INVOICE_NON_PO')
            header_data: Header fields extracted from document
            line_items: Line items extracted from document
            output_path: Optional path to save file
            
        Returns:
            Excel file bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            template = self.templates_by_type.get(doc_type)
            if not template:
                raise ValueError(f"No template found for document type: {doc_type}")
            
            excel_template = template.get('excel_template', {})
            if not excel_template:
                raise ValueError(f"No Excel template defined for document type: {doc_type}")
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_config in excel_template.get('sheets', []):
                ws = wb.create_sheet(sheet_config['name'])
                
                ws['A1'] = f"ARIA ERP - SAP {template['name']} Export"
                ws['A1'].font = Font(bold=True, size=14, color="4472C4")
                ws.merge_cells(f"A1:{self._get_column_letter(len(sheet_config['columns']))}1")
                
                ws['A2'] = f"Transaction Code: {template.get('sap_tcode', 'N/A')}"
                ws['A2'].font = Font(bold=True, size=11)
                ws.merge_cells(f"A2:{self._get_column_letter(len(sheet_config['columns']))}2")
                
                ws['A3'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A3'].font = Font(italic=True, size=10)
                ws.merge_cells(f"A3:{self._get_column_letter(len(sheet_config['columns']))}3")
                
                for col_idx, col_name in enumerate(sheet_config['columns'], 1):
                    cell = ws.cell(row=5, column=col_idx)
                    cell.value = col_name
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                if sheet_config['name'] == 'Header':
                    self._populate_header_sheet(ws, sheet_config['columns'], header_data, border)
                elif sheet_config['name'] == 'Line Items':
                    self._populate_line_items_sheet(ws, sheet_config['columns'], line_items, border)
                
                for col_idx in range(1, len(sheet_config['columns']) + 1):
                    ws.column_dimensions[self._get_column_letter(col_idx)].width = 20
            
            if output_path:
                wb.save(output_path)
                logger.info(f"Saved SAP export to {output_path}")
            
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
            
        except ImportError:
            logger.error("openpyxl not available - install openpyxl for Excel export")
            raise
        except Exception as e:
            logger.error(f"Failed to generate Excel export: {e}")
            raise
    
    def _populate_header_sheet(
        self,
        ws,
        columns: List[str],
        header_data: Dict[str, Any],
        border
    ):
        """Populate header sheet with data"""
        row_idx = 6
        
        for col_idx, col_name in enumerate(columns, 1):
            data_key = col_name.lower().replace(' ', '_').replace('-', '_')
            value = header_data.get(data_key, '')
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
    
    def _populate_line_items_sheet(
        self,
        ws,
        columns: List[str],
        line_items: List[Dict[str, Any]],
        border
    ):
        """Populate line items sheet with data"""
        for row_idx, line_item in enumerate(line_items, 6):
            for col_idx, col_name in enumerate(columns, 1):
                data_key = col_name.lower().replace(' ', '_').replace('-', '_')
                
                if data_key == 'line_no':
                    value = row_idx - 5
                else:
                    value = line_item.get(data_key, '')
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    def _get_column_letter(self, col_idx: int) -> str:
        """Convert column index to Excel column letter"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_idx)
    
    def generate_csv_export(
        self,
        doc_type: str,
        header_data: Dict[str, Any],
        line_items: List[Dict[str, Any]],
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate CSV export file for SAP upload
        
        Args:
            doc_type: SAP document type
            header_data: Header fields
            line_items: Line items
            output_path: Optional path to save file
            
        Returns:
            CSV content as string
        """
        import csv
        from io import StringIO
        
        template = self.templates_by_type.get(doc_type)
        if not template:
            raise ValueError(f"No template found for document type: {doc_type}")
        
        excel_template = template.get('excel_template', {})
        if not excel_template:
            raise ValueError(f"No Excel template defined for document type: {doc_type}")
        
        output = StringIO()
        
        for sheet_config in excel_template.get('sheets', []):
            output.write(f"# {sheet_config['name']}\n")
            
            writer = csv.DictWriter(output, fieldnames=sheet_config['columns'])
            writer.writeheader()
            
            if sheet_config['name'] == 'Header':
                row = {}
                for col_name in sheet_config['columns']:
                    data_key = col_name.lower().replace(' ', '_').replace('-', '_')
                    row[col_name] = header_data.get(data_key, '')
                writer.writerow(row)
            
            elif sheet_config['name'] == 'Line Items':
                for idx, line_item in enumerate(line_items, 1):
                    row = {}
                    for col_name in sheet_config['columns']:
                        data_key = col_name.lower().replace(' ', '_').replace('-', '_')
                        if data_key == 'line_no':
                            row[col_name] = idx
                        else:
                            row[col_name] = line_item.get(data_key, '')
                    writer.writerow(row)
            
            output.write('\n')
        
        csv_content = output.getvalue()
        
        if output_path:
            with open(output_path, 'w') as f:
                f.write(csv_content)
            logger.info(f"Saved SAP CSV export to {output_path}")
        
        return csv_content
    
    def get_template_by_type(self, doc_type: str) -> Optional[Dict]:
        """Get template by document type"""
        return self.templates_by_type.get(doc_type)
    
    def get_all_templates(self) -> List[Dict]:
        """Get all SAP export templates"""
        return self.templates
    
    def get_templates_by_module(self, module: str) -> List[Dict]:
        """Get templates for a specific SAP module"""
        return [t for t in self.templates if module.upper() in t.get('module', '').upper()]


sap_export_service = SAPExportService()
