"""
OCR and Document Classification Service for Ask Aria
"""
import os
import uuid
import hashlib
from typing import Dict, Any, Optional, List
import logging
from pathlib import Path
import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)


class OCRService:
    """OCR and document classification service"""
    
    def __init__(self, db_connection_string: str, storage_path: str = "/var/www/aria/uploads"):
        self.db_connection_string = db_connection_string
        self.storage_path = storage_path
        os.makedirs(storage_path, exist_ok=True)
    
    def get_connection(self):
        """Get database connection"""
        return psycopg2.connect(self.db_connection_string)
    
    def upload_document(
        self,
        company_id: str,
        user_id: str,
        filename: str,
        file_content: bytes,
        mime_type: str
    ) -> str:
        """
        Upload a document and store metadata
        
        Returns:
            document_id
        """
        try:
            sha256_hash = hashlib.sha256(file_content).hexdigest()
            
            document_id = str(uuid.uuid4())
            file_extension = Path(filename).suffix
            storage_filename = f"{document_id}{file_extension}"
            storage_file_path = os.path.join(self.storage_path, storage_filename)
            
            with open(storage_file_path, 'wb') as f:
                f.write(file_content)
            
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute("""
                        INSERT INTO aria_documents
                        (id, company_id, filename, mime_type, storage_path, sha256, size_bytes, uploaded_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        document_id,
                        company_id,
                        filename,
                        mime_type,
                        storage_file_path,
                        sha256_hash,
                        len(file_content),
                        user_id
                    ))
                    
                    conn.commit()
                    
                    logger.info(f"Uploaded document {document_id}: {filename}")
                    return document_id
                    
        except Exception as e:
            logger.error(f"Failed to upload document: {str(e)}")
            raise
    
    def extract_text_tesseract(self, file_path: str) -> str:
        """
        Extract text from document using Tesseract OCR or direct text extraction
        
        Args:
            file_path: Path to the document file
            
        Returns:
            Extracted text
        """
        file_extension = Path(file_path).suffix.lower()
        
        if file_extension == '.txt':
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except Exception as e:
                logger.error(f"Failed to read text file: {str(e)}")
                return ""
        
        elif file_extension in ['.xlsx', '.xls']:
            return self._extract_text_from_excel(file_path)
        
        elif file_extension == '.pdf':
            return self._extract_text_from_pdf(file_path)
        
        else:
            return self._extract_text_from_image(file_path)
    
    def _extract_text_from_excel(self, file_path: str) -> str:
        """Extract text from Excel files"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}\n")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            
            return "\n".join(text_parts)
            
        except ImportError:
            logger.warning("openpyxl not available - install openpyxl for Excel support")
            return ""
        except Exception as e:
            logger.error(f"Excel extraction failed: {str(e)}")
            return ""
    
    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF files"""
        try:
            import pytesseract
            from PIL import Image
            import pdf2image
            
            images = pdf2image.convert_from_path(file_path)
            text_parts = []
            for image in images:
                text = pytesseract.image_to_string(image)
                text_parts.append(text)
            return "\n\n".join(text_parts)
                
        except ImportError:
            logger.warning("Tesseract OCR not available - install pytesseract, pillow, pdf2image")
            return ""
        except Exception as e:
            logger.error(f"PDF OCR extraction failed: {str(e)}")
            return ""
    
    def _extract_text_from_image(self, file_path: str) -> str:
        """Extract text from image files"""
        try:
            import pytesseract
            from PIL import Image
            
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text
                
        except ImportError:
            logger.warning("Tesseract OCR not available - install pytesseract, pillow")
            return ""
        except Exception as e:
            logger.error(f"Image OCR extraction failed: {str(e)}")
            return ""
    
    def classify_document(self, document_id: str) -> Dict[str, Any]:
        """
        Classify a document based on its content
        
        Args:
            document_id: UUID of the document
            
        Returns:
            Classification result
        """
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute("""
                        SELECT * FROM aria_documents WHERE id = %s
                    """, (document_id,))
                    
                    document = cur.fetchone()
                    if not document:
                        raise ValueError(f"Document {document_id} not found")
                    
                    text = self.extract_text_tesseract(document['storage_path'])
                    
                    doc_class, confidence, labels = self._classify_text(text)
                    
                    cur.execute("""
                        INSERT INTO aria_document_classification
                        (document_id, class, confidence, labels, model)
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        document_id,
                        doc_class,
                        confidence,
                        psycopg2.extras.Json(labels),
                        "heuristic_v1"
                    ))
                    
                    conn.commit()
                    
                    logger.info(f"Classified document {document_id} as {doc_class}")
                    
                    return {
                        "document_id": document_id,
                        "class": doc_class,
                        "confidence": confidence,
                        "labels": labels,
                        "extracted_text": text[:500]
                    }
                    
        except Exception as e:
            logger.error(f"Failed to classify document: {str(e)}")
            raise
    
    def _classify_text(self, text: str) -> tuple:
        """
        Simple heuristic-based document classification
        
        Returns:
            (class, confidence, labels)
        """
        text_lower = text.lower()
        
        keywords = {
            "invoice": ["invoice", "bill", "amount due", "payment terms", "invoice number"],
            "quote": ["quotation", "quote", "proposal", "valid until", "quote number"],
            "purchase_order": ["purchase order", "po number", "delivery date", "supplier"],
            "delivery_note": ["delivery note", "delivery", "shipped", "tracking"],
            "receipt": ["receipt", "paid", "transaction", "payment received"],
            "contract": ["contract", "agreement", "terms and conditions", "parties"],
            "statement": ["statement", "account statement", "balance", "transactions"]
        }
        
        scores = {}
        for doc_type, terms in keywords.items():
            score = sum(1 for term in terms if term in text_lower)
            if score > 0:
                scores[doc_type] = score
        
        if not scores:
            return ("unknown", 0.3, [])
        
        best_class = max(scores, key=scores.get)
        max_score = scores[best_class]
        confidence = min(0.5 + (max_score * 0.1), 0.95)
        
        labels = [{"type": k, "score": v} for k, v in sorted(scores.items(), key=lambda x: x[1], reverse=True)]
        
        return (best_class, confidence, labels)
    
    def extract_fields(self, document_id: str) -> Dict[str, Any]:
        """
        Extract structured fields from a document
        
        Args:
            document_id: UUID of the document
            
        Returns:
            Extracted fields
        """
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute("""
                        SELECT d.*, c.class
                        FROM aria_documents d
                        LEFT JOIN aria_document_classification c ON d.id = c.document_id
                        WHERE d.id = %s
                    """, (document_id,))
                    
                    document = cur.fetchone()
                    if not document:
                        raise ValueError(f"Document {document_id} not found")
                    
                    text = self.extract_text_tesseract(document['storage_path'])
                    
                    fields = self._extract_fields_from_text(text, document.get('class', 'unknown'))
                    
                    cur.execute("""
                        INSERT INTO aria_document_extractions
                        (document_id, fields, confidence, model)
                        VALUES (%s, %s, %s, %s)
                        RETURNING id
                    """, (
                        document_id,
                        psycopg2.extras.Json(fields),
                        0.7,
                        "heuristic_v1"
                    ))
                    
                    conn.commit()
                    
                    logger.info(f"Extracted fields from document {document_id}")
                    
                    return {
                        "document_id": document_id,
                        "fields": fields
                    }
                    
        except Exception as e:
            logger.error(f"Failed to extract fields: {str(e)}")
            raise
    
    def _extract_fields_from_text(self, text: str, doc_class: str) -> Dict[str, Any]:
        """
        Extract structured fields from text based on document class
        
        Returns:
            Dictionary of extracted fields
        """
        import re
        
        fields = {}
        
        date_pattern = r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b'
        dates = re.findall(date_pattern, text)
        if dates:
            fields['date'] = dates[0]
        
        amount_pattern = r'[\$€£R]\s*[\d,]+\.?\d*'
        amounts = re.findall(amount_pattern, text)
        if amounts:
            fields['amount'] = amounts[0]
        
        if doc_class == 'invoice':
            invoice_num_pattern = r'invoice\s*#?\s*:?\s*([A-Z0-9-]+)'
            match = re.search(invoice_num_pattern, text, re.IGNORECASE)
            if match:
                fields['invoice_number'] = match.group(1)
        
        elif doc_class == 'quote':
            quote_num_pattern = r'quote\s*#?\s*:?\s*([A-Z0-9-]+)'
            match = re.search(quote_num_pattern, text, re.IGNORECASE)
            if match:
                fields['quote_number'] = match.group(1)
        
        elif doc_class == 'purchase_order':
            po_num_pattern = r'po\s*#?\s*:?\s*([A-Z0-9-]+)'
            match = re.search(po_num_pattern, text, re.IGNORECASE)
            if match:
                fields['po_number'] = match.group(1)
        
        return fields
